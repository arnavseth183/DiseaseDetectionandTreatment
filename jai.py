import streamlit as st
import pandas as pd
import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

st.set_page_config(page_title="Symptom Matrix Generator", layout="wide")
st.title("🦠 Disease-Symptom Binary Matrix Generator")

uploaded_file = st.file_uploader("📤 Upload your CSV or Excel file", type=["csv", "xlsx"])

if uploaded_file:
    try:
        # Load file
        if uploaded_file.name.endswith(".csv"):
            df = pd.read_csv(uploaded_file)
        else:
            df = pd.read_excel(uploaded_file, engine="openpyxl")

        # Drop S.No if exists
        df = df.drop(columns=[col for col in df.columns if col.lower() in ['s.no', 'sno', 'serial', 'id']], errors='ignore')

        # Ensure column names
        expected = {'Disease', 'Symptoms', 'Treatment', 'Medicine'}
        if not expected.issubset(set(df.columns)):
            st.error(f"Missing one or more required columns: {expected}")
            st.stop()

        # Normalize & split symptoms
        df['Symptoms'] = df['Symptoms'].fillna('').astype(str)
        df['Symptoms'] = df['Symptoms'].str.lower().str.strip()

        # Extract unique symptoms
        all_symptoms = set()
        for symptoms in df['Symptoms']:
            all_symptoms.update([s.strip() for s in symptoms.split(',') if s.strip()])

        all_symptoms = sorted(all_symptoms)

        # Initialize binary symptom matrix
        for symptom in all_symptoms:
            df[symptom] = df['Symptoms'].apply(lambda x: 1 if symptom in x else 0)

        # Final dataframe
        final_df = df.drop(columns=['Symptoms'])

        st.success("✅ Matrix created successfully. Preview below:")
        st.dataframe(final_df, use_container_width=True)

        # Save to Excel with highlights
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            final_df.to_excel(writer, index=False, sheet_name="SymptomMatrix")
            workbook = writer.book
            worksheet = writer.sheets["SymptomMatrix"]
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            for row in worksheet.iter_rows(min_row=2, min_col=4, max_col=3 + len(all_symptoms)):
                for cell in row:
                    if cell.value == 1:
                        cell.fill = yellow_fill

        output.seek(0)
        st.download_button(
            label="📥 Download Excel with Highlighted Symptoms",
            data=output,
            file_name="disease_symptom_matrix.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        st.error(f"❌ Error: {e}")



